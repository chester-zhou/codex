from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook


@dataclass
class AuditRule:
    major: str
    minor: str
    requirement: str
    standard: str


HEADER_ALIASES = {
    "major": {"大类"},
    "minor": {"小类"},
    "requirement": {"审核要求", "具体要求"},
    "standard": {"合格标准", "布局图要求项目", "备注（型号差异）"},
}


def _normalize_header(value: Optional[str]) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header_row(rows: Iterable[List[str]]) -> Optional[Dict[str, int]]:
    for idx, row in enumerate(rows, start=1):
        col_map: Dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            header = _normalize_header(cell)
            for key, aliases in HEADER_ALIASES.items():
                if header in aliases and key not in col_map:
                    col_map[key] = col_idx
        if {"major", "minor", "requirement"}.issubset(col_map.keys()):
            return {"row": idx, **col_map}
    return None


def load_rules_from_excel(path: str, sheet_name: Optional[str] = None) -> List[AuditRule]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = []
    for row in ws.iter_rows():
        rows.append([cell.value for cell in row])

    header_info = _find_header_row(rows)
    if not header_info:
        raise ValueError("未找到包含 '大类/小类/审核要求(或具体要求)' 的表头行")

    header_row = header_info["row"]
    idx_major = header_info["major"]
    idx_minor = header_info["minor"]
    idx_req = header_info["requirement"]
    idx_std = header_info.get("standard")

    rules: List[AuditRule] = []
    current_major = ""

    for row in rows[header_row:]:
        major = _normalize_header(row[idx_major]) if idx_major < len(row) else ""
        minor = _normalize_header(row[idx_minor]) if idx_minor < len(row) else ""
        requirement = _normalize_header(row[idx_req]) if idx_req < len(row) else ""
        standard = _normalize_header(row[idx_std]) if idx_std is not None and idx_std < len(row) else ""

        if not any([major, minor, requirement, standard]):
            continue

        if major:
            current_major = major
        else:
            major = current_major

        if not requirement:
            continue

        rules.append(
            AuditRule(
                major=major,
                minor=minor,
                requirement=requirement,
                standard=standard,
            )
        )

    if not rules:
        raise ValueError("未解析到任何审核要点")

    return rules
